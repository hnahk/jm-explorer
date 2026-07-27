#!/usr/bin/env python3
"""
JM Explorer — data pipeline (the "backend").

Encodes the JoyMade relationship graph and joins it into one traceable dataset.

  FINANCIAL ─(occasion / evergreen)─ PLAN
  PLAN ─(occasion·niche·recipient·persona ; unique = IDEA CODE)─ IDEA
  IDEA ─(product live link)─ DESIGN & LISTING
  IDEA ─(idea code + product link, 1→many)─ MEDIA (videos, each with a TB##### id)
  MEDIA ─(TB##### numbering id embedded in campaign name)─ ADS
  ADS ─(rev·profit·roas·cpa·cpm·[cpc])─ FINANCIAL

Universal hard key across code systems = the design number  C####/B####.

Output:  ../data.js   (window.JM = {...}; loaded by explorer.html via <script>)
         ../data.json (same payload)

Refresh raw API data via backend/welast.py (cached in raw/ so this build is offline):
  /market/products/store/20 -> raw/products.json   (catalog: handle, tags, shopifyPublishedAt, type)
  /report/stores/campaigns/5 -> raw/campaigns.json  (per-campaign spend/rev/orders/impressions)
  NOTE: add 'inlineLinkClicks' to the campaign field list to enable CPC.
"""
import json, re, os, glob, calendar
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
RAW  = os.path.join(HERE, "raw")
OUT_JS   = os.path.join(HERE, "..", "data.js")
OUT_JSON = os.path.join(HERE, "..", "data.json")
DL = os.path.expanduser("~/Downloads")
prod_matches = glob.glob(os.path.join(DL, "idea*.xlsx")) + glob.glob(os.path.join(DL, "WorkFlows*Product Table*.xlsx"))
PROD_XLSX = max(prod_matches, key=os.path.getmtime) if prod_matches else None
media_matches = glob.glob(os.path.join(DL, "media*.xlsx")) + glob.glob(os.path.join(DL, "WorkFlows*Media Table*.xlsx"))
MEDIA_XLSX = max(media_matches, key=os.path.getmtime) if media_matches else None
PLAN_XLSX  = os.path.join(DL, "JM Idea's Plan .xlsx")

# ---------- helpers ----------
def dcode(s):
    m = re.search(r'[cbCB]\d{3,5}', str(s) or ''); return m.group().upper() if m else None
def f(x):
    try: return float(x or 0)
    except: return 0.0
def iso(d):
    return d.isoformat()[:10] if isinstance(d, (date, datetime)) else None
def asdate(v):
    return v.date() if isinstance(v, datetime) else (v if isinstance(v, date) else None)
def days(a, b):
    return (b - a).days if (a and b) else None
def namedate(nm):
    best = None
    for mo in re.finditer(r'20(\d{2})(\d{2})(\d{2})', nm or ''):
        y, mm, dd = 2000 + int(mo.group(1)), int(mo.group(2)), int(mo.group(3))
        if 2024 <= y <= 2027 and 1 <= mm <= 12 and 1 <= dd <= 31:
            try:
                d = date(y, mm, dd); best = d if (best is None or d < best) else best
            except: pass
    return best
def occ_canon(s):
    raw = str(s or '').strip()
    if not raw or raw.lower() == 'none': return 'Unspecified'
    t = raw.lower()
    for k, v in [('valentine','Valentine'),('mother',"Mother's Day"),('father',"Father's Day"),
                 ('graduation','Graduation'),('back to school','Back to School'),('teacher','Back to School'),
                 ('250','4th/Patriotic'),('independence','4th/Patriotic'),('patriot','4th/Patriotic'),('memorial day','4th/Patriotic'),
                 ('world cup','World Cup'),('soccer','World Cup'),('christmas','Christmas'),('easter','Easter'),('halloween','Halloween'),
                 ('anniversary','Anniversary'),('wedding','Wedding'),('summer','Summer'),('birthday','Birthday'),
                 ('newborn','Newborn'),('baby','Newborn'),('pregnan','Newborn'),('retirement','Retirement'),
                 ('nurse','Nurses Day'),('best friend','Best Friends'),('black history','Black History'),
                 ('book','Book Lovers'),('library','Book Lovers'),('evergreen','Evergreen')]:
        if k in t: return v
    return raw.split(',')[0].strip()[:24]   # preserve the team's actual occasion instead of collapsing to Other
OCC_DATE = {'Valentine':'2026-02-14','Easter':'2026-04-05',"Mother's Day":'2026-05-10',
            "Father's Day":'2026-06-21','Graduation':'2026-05-20','4th/Patriotic':'2026-07-04',
            'World Cup':'2026-06-11','Christmas':'2026-12-25','Halloween':'2026-10-31','Back to School':'2026-08-15'}

def canon_rec(s):
    """Crosswalk both the plan's niches and Product Table's Recipient into one shared recipient taxonomy
       (the plan niches map to the Recipient field, not the coarse 'Niche' field which collapses to 'Family')."""
    t = str(s or '').lower()
    if not t or t == 'none': return 'Unspecified'
    if 'reaction' in t or 'scale' in t: return 'Reaction/Scale'
    if any(k in t for k in ['husband','wife','couple','vợ','chồng',' vk',' ck']): return 'Couple'
    if 'memorial' in t or 'tưởng' in t or 'loving memory' in t or 'sympathy' in t: return 'Memorial'
    if 'mother-in-law' in t or 'mom' in t or 'mother' in t or 'mẹ' in t: return 'Mom'
    if 'dad' in t or 'father' in t or 'bố' in t: return 'Dad'
    if 'grandma' in t or 'grandpa' in t or 'grandparent' in t: return 'Grandparents'
    if 'granddaughter' in t or 'grandson' in t or 'grandkid' in t: return 'Grandkids'
    if 'daughter' in t or 'son' in t or 'kid' in t or 'child' in t or 'baby' in t or 'first parent' in t or 'pregnan' in t: return 'Kids'
    if 'sister' in t or 'brother' in t or 'sibling' in t: return 'Siblings'
    if 'besti' in t or 'friend' in t or 'coworker' in t: return 'Besties'
    if 'selfbuy' in t or 'self' in t: return 'Self-buy'
    if 'pet' in t or 'fur ' in t: return 'Pet'
    if 'teacher' in t or 'nurse' in t: return 'Teacher/Nurse'
    if 'family' in t: return 'Family (unsplit)'
    return 'Other'

# ---------- 1. PLAN ----------
def num(v):
    try: return int(float(v))
    except: return None
plan_weeks = {}
plan_rec_planned = defaultdict(int)   # recipient bucket -> planned idea qty (across all weeks)
cache_plan_file = os.path.join(RAW, "cache_plan.json")

if os.path.exists(PLAN_XLSX):
    wb = load_workbook(PLAN_XLSX, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]; hr = qi = ci = ni = None
        for r in range(1, 4):
            row = [str(ws.cell(r, c).value or '').lower() for c in range(1, 15)]
            if any('idea quantity' in v for v in row):
                hr = r; qi = next(i + 1 for i, v in enumerate(row) if 'idea quantity' in v)
                ci = next((i + 1 for i, v in enumerate(row) if 'created' in v), None)
                ni = next((i + 1 for i, v in enumerate(row) if 'niche' in v), None); break
        if not qi: continue
        theme = ''
        for c in range(1, 15):
            v = str(ws.cell(hr, c).value or '')
            if any(w in v.lower() for w in ('evergreen','valentine','easter','mother','father','christmas','hobbies','skip','graduation','patriot','independence','summer','world cup')):
                theme = v.replace('\n', ' ').strip()[:60]; break
        wp = wm = 0
        for r in range(hr + 1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or '').strip().upper().startswith('TOTAL'): continue
            q = num(ws.cell(r, qi).value); m = num(ws.cell(r, ci).value) if ci else None
            wp += q or 0; wm += m or 0
            if q and ni: plan_rec_planned[canon_rec(ws.cell(r, ni).value)] += q
        plan_weeks[sn] = {'planned': wp, 'created': wm, 'theme': theme}
    wb.close()
    with open(cache_plan_file, "w") as fp:
        json.dump({'plan_weeks': plan_weeks, 'plan_rec_planned': dict(plan_rec_planned)}, fp)
elif os.path.exists(cache_plan_file):
    cd = json.load(open(cache_plan_file))
    plan_weeks = cd.get('plan_weeks', {})
    for k, v in cd.get('plan_rec_planned', {}).items():
        plan_rec_planned[k] = v

def _mkrange(sn):
    m = re.match(r'(\d+)W(\d+)(?:.*?W?(\d+))?', sn)
    if not m: return None
    mo = int(m.group(1)); w1 = int(m.group(2)); w2 = int(m.group(3)) if m.group(3) else w1
    yr = 2025 if mo >= 8 else 2026
    try:
        ld = calendar.monthrange(yr, mo)[1]
        return (date(yr, mo, min((w1 - 1) * 7 + 1, ld)), date(yr, mo, min(w2 * 7, ld)))
    except Exception: return None
plan_ranges = [(sn, _mkrange(sn)[0], _mkrange(sn)[1]) for sn in plan_weeks if _mkrange(sn)]
def match_week(d):
    if not d: return None
    for sn, st, en in plan_ranges:
        if st <= d <= en: return sn
    best, bd = None, 10**9
    for sn, st, en in plan_ranges:
        mid = st + (en - st) / 2; dd = abs((mid - d).days)
        if dd < bd: bd, best = dd, sn
    return best if bd <= 21 else None

# ---------- 2. IDEA + DESIGN (Product Table) ----------
prod = {}
cache_prod_file = os.path.join(RAW, "cache_prod.json")
if PROD_XLSX and os.path.exists(PROD_XLSX):
    wb = load_workbook(PROD_XLSX, data_only=True); ws = wb[wb.sheetnames[0]]
    for r in range(2, ws.max_row + 1):
        c = dcode(ws.cell(r, 2).value)
        if not c: continue
        prod.setdefault(c, {
            'idea_name': ws.cell(r, 1).value, 'creator': ws.cell(r, 9).value,
            'full_code': (str(ws.cell(r, 2).value or '')).strip(),   # B: full Idea Code, e.g. JMKSBTI4C6993
            'occ': occ_canon(ws.cell(r, 19).value), 'niche': ws.cell(r, 20).value,
            'recipient': ws.cell(r, 21).value, 'persona': ws.cell(r, 22).value, 'product': ws.cell(r, 8).value,
            'idea': iso(asdate(ws.cell(r, 30).value)), 'design': iso(asdate(ws.cell(r, 46).value))})
    wb.close()
    with open(cache_prod_file, "w") as fp:
        json.dump(prod, fp)
elif os.path.exists(cache_prod_file):
    prod = json.load(open(cache_prod_file))

# Convert date strings back to date objects for internal calculation
for c, p in prod.items():
    if isinstance(p.get('idea'), str):
        try: p['idea'] = date.fromisoformat(p['idea'])
        except: pass
    if isinstance(p.get('design'), str):
        try: p['design'] = date.fromisoformat(p['design'])
        except: pass

# ---------- 3. LISTING (catalog) ----------
cat = json.load(open(os.path.join(RAW, "products.json")))
catc = {}
for p in cat:
    c = dcode(p.get('handle'))
    if not c: continue
    hf = re.search(r'jm[a-z0-9]+[cb]\d{3,5}', (p.get('handle') or '').lower())
    catc.setdefault(c, {'live': p['shopifyPublishedAt'][:10] if p.get('shopifyPublishedAt') else None,
                        'type': p.get('marketProductType'),
                        'link': 'https://joymade.co/products/' + (p.get('handle') or ''),
                        'full': hf.group().upper() if hf else None,
                        'occ': occ_canon(p.get('tags')),
                        'title': p.get('title')})

# ---------- 4. MEDIA (1 idea -> many videos) ----------
media = defaultdict(list)   # code -> [ {date, video, tb, link} ]
cache_media_file = os.path.join(RAW, "cache_media.json")
if MEDIA_XLSX and os.path.exists(MEDIA_XLSX):
    wb = load_workbook(MEDIA_XLSX, data_only=True); ws = wb[wb.sheetnames[0]]
    for r in range(2, ws.max_row + 1):
        c = dcode(ws.cell(r, 1).value) or dcode(ws.cell(r, 40).value)      # A idea code OR AN parent (video sub-rows)
        if not c: continue
        tb = ws.cell(r, 44).value                                          # AR Numbering ID  (media<->ads key)
        row = {'date': iso(asdate(ws.cell(r, 26).value) or asdate(ws.cell(r, 25).value)),  # Z End / Y Shooting
               'video': (ws.cell(r, 2).value or ''),                       # B Video Name
               'tb': (str(tb).strip() if tb else ''), 'vlink': (ws.cell(r, 16).value or ''),
               'link': (ws.cell(r, 51).value or '')}                       # AY Product Link
        if row['tb'] or row['date'] or row['video'] or row['vlink']: media[c].append(row)
    wb.close()
    with open(cache_media_file, "w") as fp:
        json.dump(dict(media), fp)
elif os.path.exists(cache_media_file):
    cd = json.load(open(cache_media_file))
    for k, v in cd.items():
        media[k] = v

media_first = {c: min((v['date'] for v in vs if v['date']), default=None) for c, vs in media.items()}
# reverse index: TB numbering id -> media video (to link a campaign to its video)
tb_to_media = {v['tb']: {**v, 'code': c} for c, vs in media.items() for v in vs if v['tb']}

# ---------- 5. ADS + FINANCIAL (campaigns) ----------
camps = json.load(open(os.path.join(RAW, "campaigns.json")))
cagg = defaultdict(lambda: [0.0, 0.0, 0])        # code -> [spend, rev, orders]
adsfirst = {}
camp_by_code = defaultdict(list)
seen = {}; camp_full = {}
for m, cs in camps.items():
    for c in cs:
        nm = c['name'] or ''
        mm = re.search(r'JM[A-Z0-9]{4,}[CB]\d{3,5}', nm); cc = dcode(mm.group()) if mm else None
        if cc and mm: camp_full.setdefault(cc, mm.group())
        sp, rv, od, im, cl = f(c['spend']), f(c['rev']), int(f(c['orderCount'])), int(f(c.get('impressions'))), int(f(c.get('inlineLinkClicks')))
        if cc:
            a = cagg[cc]; a[0] += sp; a[1] += rv; a[2] += od
            d = namedate(nm)
            if d and (cc not in adsfirst or d < adsfirst[cc]): adsfirst[cc] = d
        uid = c.get('id') or nm
        if uid in seen:                                   # same campaign, another month -> merge
            row = seen[uid]; row['spend'] += sp; row['rev'] += rv; row['orders'] += od; row['impr'] += im; row['clk'] += cl; continue
        if not cc: continue
        tbm = re.search(r'TB\d+', nm)                     # media<->ads join key
        row = {'name': nm[:90], 'date': iso(namedate(nm)), 'media_id': tbm.group() if tbm else '',
               'spend': sp, 'rev': rv, 'orders': od, 'impr': im, 'clk': cl, 'fb': str(c.get('fbId') or ''), 'thumb': c.get('thumbnail') or ''}
        seen[uid] = row; camp_by_code[cc].append(row)

# ---------- 5.b. SHOPIFY ORDERS (true net revenue & FBT) ----------
shop_stats = defaultdict(lambda: {'orders': 0, 'net_rev': 0.0, 'items': defaultdict(int)})
if os.path.exists(os.path.join(RAW, "orders.json")):
    try:
        orders_data = json.load(open(os.path.join(RAW, "orders.json")))
        for m, olist in orders_data.items():
            for o in olist:
                dcodes = [dcode(c) for c in o.get('designCodes', [])]
                dcodes = [c for c in dcodes if c]
                net = float(o.get('totalPrice', 0)) - float(o.get('cogs', 0))
                for c in set(dcodes):
                    st = shop_stats[c]
                    st['orders'] += 1
                    st['net_rev'] += net
                    for it in o.get('items', []):
                        st['items'][it] += 1
    except Exception as e:
        print("Error parsing orders.json:", e)

# ---------- assemble per-design records (the trace hub) ----------
BREAKEVEN = 1.43
def bucket(sp, rv, od):
    if sp <= 0: return 'not-advertised'
    if od < 20: return 'untestable'
    return 'winner' if rv / sp >= BREAKEVEN else 'loser'

codes = set(cagg) | set(prod) | set(catc)
designs = []
for c in sorted(codes):
    p = prod.get(c, {}); ct = catc.get(c, {}); vids = media.get(c, [])
    sp, rv, od = cagg.get(c, [0.0, 0.0, 0]); idea = p.get('idea'); ads = adsfirst.get(c)
    _cs = camp_by_code.get(c, [])
    _img = [r['date'] for r in _cs if not r.get('media_id') and r['date']]   # no TB video -> image ad
    _vid = [r['date'] for r in _cs if r.get('media_id') and r['date']]       # carries a TB video -> video ad
    occ = p.get('occ') or ct.get('occ') or 'Unspecified'
    live_d = date.fromisoformat(ct['live']) if ct.get('live') else None
    pw = match_week(idea); pv = 'idea'
    if not pw: pw, pv = match_week(live_d), 'live'
    if not pw: pw, pv = match_week(ads), 'ads'
    if not pw: pv = None
    
    # Shopify Stats
    sst = shop_stats.get(c, {'orders': 0, 'net_rev': 0.0, 'items': {}})
    shop_orders = sst['orders']
    shop_net_rev = sst['net_rev']
    shop_net_aov = shop_net_rev / shop_orders if shop_orders > 0 else 0
    # Top 3 FBT items (excluding the design's own main product if possible, but we'll just take top 3)
    fbt = [k for k, v in sorted(sst['items'].items(), key=lambda x: -x[1])[:3]]

    designs.append({
        'code': c, 'has_idea': c in prod, 'full': (prod.get(c,{}).get('full_code') or (catc.get(c,{}) or {}).get('full') or camp_full.get(c) or c), 'idea': p.get('idea_name') or ct.get('title') or '', 'creator': p.get('creator') or '',
        'occasion': occ, 'niche': p.get('niche') or '', 'recipient': p.get('recipient') or '', 'rec': canon_rec(p.get('recipient')),
        'persona': p.get('persona') or '', 'product': ct.get('type') or (p.get('product') or ''),
        'link': ct.get('link') or '', 'plan_week': pw if pw in plan_weeks else None, 'plan_via': pv,
        'd_idea': iso(idea), 'd_design': iso(p.get('design')), 'd_live': ct.get('live'),
        'd_media': media_first.get(c), 'd_ads': iso(ads), 'd_ad_img': (min(_img) if _img else None), 'd_ad_vid': (min(_vid) if _vid else None), 'nvideos': len(vids), 'videos': vids,
        'lat_idea_live': days(idea, date.fromisoformat(ct['live'])) if (idea and ct.get('live')) else None,
        'lat_idea_ads': days(idea, ads), 'occ_date': OCC_DATE.get(occ),
        'days_before_occ': (date.fromisoformat(OCC_DATE[occ]) - ads).days if (occ in OCC_DATE and ads) else None,
        'spend': round(sp, 2), 'rev': round(rv, 2), 'profit': round(0.70 * rv - sp, 2),
        'roas': round(rv / sp, 3) if sp else 0, 'orders': od,
        'cpa': round(sp / od, 2) if od else None, 'ncamp': len(camp_by_code.get(c, [])), 'ncamp_img': sum(1 for r in _cs if not r.get('media_id')), 'ncamp_vid': sum(1 for r in _cs if r.get('media_id')),
        'bucket': bucket(sp, rv, od),
        'shop_orders': shop_orders, 'shop_net_rev': round(shop_net_rev, 2), 'shop_net_aov': round(shop_net_aov, 2), 'fbt': fbt
    })

# keep only records whose listing went live on/after 1 Dec 2025 (align to the plan window; drop older products)
CUTOFF = '2025-12-01'
designs = [d for d in designs if (not d['d_live']) or d['d_live'] >= CUTOFF]
_keep = set(d['code'] for d in designs)
campaigns = {c: rows for c, rows in camp_by_code.items() if c in _keep}

# ---------- aggregates ----------
adv = [d for d in designs if d['spend'] > 0]
tot_spend = round(sum(d['spend'] for d in adv)); tot_rev = round(sum(d['rev'] for d in adv))
b = {'winner': [0, 0.0, 0.0], 'loser': [0, 0.0, 0.0], 'untestable': [0, 0.0, 0.0]}
for d in adv:
    x = b[d['bucket']]; x[0] += 1; x[1] += d['spend']; x[2] += d['profit']
summary = {'designs': len(designs), 'advertised': len(adv), 'spend': tot_spend, 'rev': tot_rev,
           'profit': round(0.70 * tot_rev - tot_spend), 'breakeven_roas': BREAKEVEN,
           'funnel': {'created': len(designs), 'live': len([d for d in designs if d['d_live']]),
                      'advertised': len(adv), 'profitable': b['winner'][0]},
           'p0': {k: {'n': v[0], 'spend': round(v[1]), 'profit': round(v[2])} for k, v in b.items()},
           'plan_total_planned': sum(w['planned'] for w in plan_weeks.values()),
           'plan_total_created': sum(w['created'] for w in plan_weeks.values())}

# recipient plan-vs-actual: plan niches crosswalked to Recipient (canon_rec) vs actual designs' recipient
rec_agg = {}
for d in designs:
    r = rec_agg.setdefault(d['rec'], {'planned': 0, 'created': 0, 'advertised': 0, 'spend': 0.0, 'rev': 0.0})
    r['created'] += 1
    if d['spend'] > 0:
        r['advertised'] += 1; r['spend'] += d['spend']; r['rev'] += d['rev']
for k, planned in plan_rec_planned.items():
    rec_agg.setdefault(k, {'planned': 0, 'created': 0, 'advertised': 0, 'spend': 0.0, 'rev': 0.0})['planned'] += planned
recipient_plan = {k: {'planned': v['planned'], 'created': v['created'], 'advertised': v['advertised'],
                      'spend': round(v['spend']), 'profit': round(0.70 * v['rev'] - v['spend']),
                      'roas': round(v['rev'] / v['spend'], 2) if v['spend'] else 0} for k, v in rec_agg.items()}

edges = {'financial<->plan': 'occasion / evergreen',
         'plan<->idea': 'occasion · niche · recipient · persona ; unique = idea code',
         'idea<->listing': 'product live link (shopifyPublishedAt)',
         'idea<->media': 'idea code + product link (1 idea -> many videos)',
         'media<->ads': 'TB##### numbering id in campaign name',
         'ads<->financial': 'rev · profit · roas · cpa · cpm · cpc'}

payload = {'summary': summary, 'edges': edges, 'designs': designs, 'campaigns': campaigns,
           'plan_weeks': plan_weeks, 'recipient_plan': recipient_plan, 'generated_for': '2026 H1',
           'note': 'FB channel only; profit = 0.70*rev - spend; breakeven ROAS 1.43; CPA/CPM/CPC from orders/impressions/clicks'}
json.dump(payload, open(OUT_JSON, 'w'), ensure_ascii=False)
with open(OUT_JS, 'w') as fp:
    fp.write("window.JM = "); json.dump(payload, fp, ensure_ascii=False); fp.write(";")
print(f"designs={len(designs)} advertised={len(adv)} campaigns={sum(len(v) for v in campaigns.values())} media_ideas={len(media)}")
print(f"summary profit ${summary['profit']:,}  funnel {summary['funnel']}")
print(f"wrote data.js ({os.path.getsize(OUT_JS)//1024} KB)")