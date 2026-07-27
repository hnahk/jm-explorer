#!/usr/bin/env python3
"""
Top-down dashboard aggregates -> ../dash.js (window.DASH).
Four data spaces:
  1 Overview        financial -> occasion/evergreen list & % -> plan totals
  2 Plan vs Actual  detailed plan rows + per-week planned vs actually-live + 'served the plan?' rate
                    (an idea counts as executed-on-plan only if its FIRST-LIVE date falls in a planned week)
  3 Ads Efficiency  ROAS distribution, spend concentration, best/worst ideas
  4 Timing          occasion lead-time (days before the occasion the first ad launched)
"""
import json, re, os, calendar
from collections import defaultdict
from datetime import date
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.join(HERE, "..")
DL = os.path.expanduser("~/Downloads")
PLAN_XLSX = os.path.join(DL, "JM Idea's Plan .xlsx")
J = json.load(open(os.path.join(ROOT, "data.json")))
DES = J["designs"]; ADV = [d for d in DES if d["spend"] > 0]
BREAKEVEN = 1.43
OCC_DATE = {'Valentine':date(2026,2,14),'Easter':date(2026,4,5),"Mother's Day":date(2026,5,10),
            "Father's Day":date(2026,6,21),'Graduation':date(2026,5,20),'4th/Patriotic':date(2026,7,4),
            'World Cup':date(2026,6,11),'Christmas':date(2026,12,25),'Halloween':date(2026,10,31),'Back to School':date(2026,8,15)}

# ---------- 1. OVERVIEW ----------
occ = defaultdict(lambda: [0.0,0.0,0])   # occasion -> [spend, rev, n]
for d in ADV:
    o = occ[d["occasion"]]; o[0]+=d["spend"]; o[1]+=d["rev"]; o[2]+=1
tot_spend = sum(d["spend"] for d in ADV); tot_rev = sum(d["rev"] for d in ADV)
tot_loss = -sum(min(0, 0.70*v[1]-v[0]) for v in occ.values()) or 1
occasions = []
for name,(sp,rv,n) in sorted(occ.items(), key=lambda kv:-kv[1][0]):
    pr = 0.70*rv-sp
    occasions.append({"name":name,"evergreen": name=="Evergreen","spend":round(sp),"rev":round(rv),
                      "profit":round(pr),"roas":round(rv/sp,2) if sp else 0,"n":n,
                      "pct_spend":round(sp/tot_spend*100,1),"pct_loss":round(-pr/tot_loss*100,1) if pr<0 else 0})
ever_sp = sum(d["spend"] for d in ADV if d["occasion"]=="Evergreen")
ever_rv = sum(d["rev"] for d in ADV if d["occasion"]=="Evergreen")
overview = {"kpi":{"profit":round(0.70*tot_rev-tot_spend),"spend":round(tot_spend),"rev":round(tot_rev),
                   "roas":round(tot_rev/tot_spend,2),"breakeven":BREAKEVEN,"advertised":len(ADV),"designs":len(DES)},
            "occasions":occasions, "funnel":J["summary"]["funnel"], "p0":J["summary"]["p0"],
            "evergreen":{"spend":round(ever_sp),"profit":round(0.70*ever_rv-ever_sp),"roas":round(ever_rv/ever_sp,2) if ever_sp else 0,
                         "seasonal_spend":round(tot_spend-ever_sp),"seasonal_profit":round(0.70*(tot_rev-ever_rv)-(tot_spend-ever_sp))}}

# ---------- 2. PLAN vs ACTUAL ----------
def num(v):
    try: return int(float(v))
    except: return None
def mkrange(sn):
    m=re.match(r'(\d+)W(\d+)(?:.*?W?(\d+))?', sn)
    if not m: return None
    mo=int(m.group(1)); w1=int(m.group(2)); w2=int(m.group(3)) if m.group(3) else w1
    yr=2025 if mo>=8 else 2026
    try:
        ld=calendar.monthrange(yr,mo)[1]; return (date(yr,mo,min((w1-1)*7+1,ld)), date(yr,mo,min(w2*7,ld)))
    except: return None
plan_rows=[]; week_meta={}
try:
    wb=load_workbook(PLAN_XLSX, data_only=True)
except Exception as e:
    print(f"Warning: could not load plan {PLAN_XLSX}: {e}")
    class DummyWB:
        sheetnames = []
        def close(self): pass
    wb = DummyWB()
for sn in wb.sheetnames:
    ws=wb[sn]; hr=qi=ci=ni=pi=pti=pici=pendi=None
    for r in range(1,4):
        row=[str(ws.cell(r,c).value or '').lower() for c in range(1,16)]
        if any('idea quantity' in v for v in row):
            hr=r; qi=next(i+1 for i,v in enumerate(row) if 'idea quantity' in v)
            ci=next((i+1 for i,v in enumerate(row) if 'created' in v),None)
            ni=next((i+1 for i,v in enumerate(row) if 'niche' in v),None)
            pti=next((i+1 for i,v in enumerate(row) if 'product type' in v),None)
            pici=next((i+1 for i,v in enumerate(row) if v.strip()=='pic'),None)
            pendi=next((i+1 for i,v in enumerate(row) if 'pending' in v),None)
            pi=next((i+1 for i,v in enumerate(row) if v.strip()=='product'),None) or next((i+1 for i,v in enumerate(row) if 'product' in v and 'type' not in v),None)
            break
    if not qi: continue
    theme=''
    for c in range(1,16):
        v=str(ws.cell(hr,c).value or '')
        if any(w in v.lower() for w in ('evergreen','valentine','easter','mother','father','christmas','hobbies','skip','graduation','patriot','independence','summer','world cup')):
            theme=v.replace('\n',' ').strip()[:50]; break
    wp=wm=0; last_prod=''; last_pt=''; last_sec=''; seci=1 if pti else None
    for r in range(hr+1, ws.max_row+1):
        c1=str(ws.cell(r,1).value or '').strip()
        if c1.upper().startswith('TOTAL'): continue
        q=num(ws.cell(r,qi).value); cr=num(ws.cell(r,ci).value) if ci else None
        if q is None and cr is None: continue
        if seci and c1: last_sec=c1.replace(chr(10),' ').strip()[:44]
        if pti:
            pt=ws.cell(r,pti).value
            if pt and str(pt).strip(): last_pt=str(pt).strip().replace(chr(10),' ')[:26]
        prod=ws.cell(r,pi).value if pi else None
        prod=str(prod).strip().replace(chr(10),' ')[:36] if (prod and str(prod).strip()) else ''
        if prod: last_prod=prod
        niche=(str(ws.cell(r,ni).value).strip().replace(chr(10),' ')[:44] if ni and ws.cell(r,ni).value else '')
        pic=(str(ws.cell(r,pici).value).strip()[:12] if pici and ws.cell(r,pici).value else '')
        pend=(str(ws.cell(r,pendi).value).strip().replace(chr(10),' ')[:40] if pendi and ws.cell(r,pendi).value else '')
        plan_rows.append({"week":sn,"section":last_sec,"product_type":last_pt,"product":prod or last_prod,"niche":niche,"planned":q or 0,"pic":pic,"created":cr or 0,"pending":pend})
        wp+=q or 0; wm+=cr or 0
    rg=mkrange(sn)
    week_meta[sn]={"theme":theme,"planned":wp,"created":wm,"start":rg[0].isoformat() if rg else None,"end":rg[1].isoformat() if rg else None,"rg":rg}
wb.close()
# actual: designs first-live in each plan week; served flag
def dt(s): return date.fromisoformat(s) if s else None
weeks_sorted=sorted([k for k in week_meta if week_meta[k]["rg"]], key=lambda k:week_meta[k]["rg"][0])
def week_of(d):
    if not d: return None
    for sn in weeks_sorted:
        st,en=week_meta[sn]["rg"]
        if st<=d<=en: return sn
    return None
live_by_week=defaultdict(int); served=0; served_adv=0
for d in DES:
    w=week_of(dt(d.get("d_live")))
    if w:
        live_by_week[w]+=1; served+=1
        if d["spend"]>0: served_adv+=1
weeks=[]
for sn in weeks_sorted:
    m=week_meta[sn]
    weeks.append({"week":sn,"theme":m["theme"],"start":m["start"],"planned":m["planned"],
                  "created_plan":m["created"],"live_actual":live_by_week.get(sn,0),
                  "adherence":round(live_by_week.get(sn,0)/m["planned"]*100) if m["planned"] else None})
plan = {"rows":plan_rows,"weeks":weeks,
        "served":{"designs":len(DES),"served":served,"rate":round(served/len(DES)*100),
                  "advertised":len(ADV),"served_adv":served_adv,"rate_adv":round(served_adv/len(ADV)*100)},
        "plan_total_planned":sum(m["planned"] for m in week_meta.values()),
        "plan_total_created":sum(m["created"] for m in week_meta.values())}

# ---------- 2b. PLAN HIERARCHY: MASTERPLAN -> JM plan -> actual ideas ----------
from datetime import timedelta
def canon_product(x):
    t=str(x or '').lower()
    for k,v in [('whiskey','Whiskey Glass/Bottle'),('beer glass','Glass'),('mug','Mug'),('shirt','Shirt'),('t-shirt','Shirt'),('tee','Shirt'),
       ('hoodie','Hoodie'),('tumbler','Tumbler'),('plaque','Plaque'),('bunny','Bunny'),('teddy','Bunny'),('lamp','Lamp'),('pocket hug','Pocket Hug'),
       ('apron','Apron'),('onesie','Onesie'),('cap','Cap'),('wallet','Wallet'),('vent clip','Car Vent Clip'),('doormat','Doormat'),
       ('ornament','Ornament'),('suncatcher','Suncatcher'),('blanket','Blanket'),('bracelet','Bracelet'),('music box','Music Box'),
       ('watch','Watch'),('candle','Candle'),('sign','Sign'),('glass','Glass'),('acrylic','Acrylic'),('keychain','Keychain'),
       ('necklace','Jewelry'),('flag','Flag'),('pen holder','Stationery'),('charm','Jewelry')]:
        if k in t: return v
    return (str(x or 'Other').strip() or 'Other')[:20]

MASTER={}
try:
    mwb=load_workbook(os.path.join(DL,'H1.2026 MASTERPLAN  IDEA.xlsx'), data_only=True)
    for sn,occ_name,prow in [('MD',"Mother's Day",8),('FD',"Father's Day",3)]:
        ws=mwb[sn]; subs=[]
        for c in range(2,20):
            dname=str(ws.cell(2,c).value or '').strip()
            if not dname: continue
            pv=ws.cell(prow,c).value
            try: pv=float(pv); pv=round(pv*100) if pv<=1 else round(pv)
            except: pv=None
            try: sv=int(float(ws.cell(4,c).value))
            except: sv=None
            subs.append({'dir':dname[:40],'pct':pv,'searchvol':sv,'rank':str(ws.cell(5,c).value or '')[:14]})
        MASTER[occ_name]=subs
    mwb.close()
except Exception as e: print('masterplan load err', e)

def occ_window(o):
    d=OCC_DATE.get(o); return (d-timedelta(days=56), d) if d else None
JM_TIMELINE = defaultdict(list)
try:
    mwb = load_workbook(os.path.join(DL, 'H1.2026 MASTERPLAN  IDEA.xlsx'), data_only=True)
    if 'TIMELINE' in mwb.sheetnames:
        ws_t = mwb['TIMELINE']; m_str = None
        for c in range(32, 60):
            v5 = str(ws_t.cell(5, c).value or '').strip()
            if v5: m_str = v5
            v6 = str(ws_t.cell(6, c).value or '').strip()
            jm = str(ws_t.cell(15, c).value or '').strip()
            if jm and m_str and v6:
                m_match = re.search(r'(\d+)', m_str); w_match = re.search(r'(\d+)', v6)
                if m_match and w_match:
                    mo = int(m_match.group(1)); yr = 2025 if mo >= 8 else 2026
                    w1 = int(w_match.group(1)); ld = calendar.monthrange(yr,mo)[1]
                    ws_date = date(yr,mo,min((w1-1)*7+1,ld))
                    for _o in OCC_DATE:
                        w = occ_window(_o)
                        if w and w[0] <= ws_date <= w[1]:
                            JM_TIMELINE[_o].append({'week': f'{m_str} {v6}', 'text': jm})
    mwb.close()
except Exception as e: print('jm timeline load err', e)
week_start={sn:week_meta[sn]['rg'][0] for sn in week_meta if week_meta[sn]['rg']}
def row_occasions(pr):
    ws=week_start.get(pr['week']); occs=[]
    for o in OCC_DATE:
        w=occ_window(o)
        if ws and w and w[0]<=ws<=w[1]: occs.append(o)
    if 'evergreen' in week_meta.get(pr['week'],{}).get('theme','').lower(): occs.append('Evergreen')
    return occs
act_by_occ=defaultdict(list)
for d in ADV: act_by_occ[d['occasion']].append(d)
OCC_ORDER=["Valentine","Easter","Mother's Day","Graduation","Father's Day","4th/Patriotic","Back to School","World Cup","Christmas","Evergreen"]
hier=[]
for o in OCC_ORDER:
    jm_plan=[]
    for pr in plan_rows:
        if o in row_occasions(pr) and (pr['planned'] or pr['product']):
            cp=canon_product(pr['product']); wdw=occ_window(o) if o in OCC_DATE else None
            wm_match=re.match(r'(\d+)W', pr['week'])
            p_mo=int(wm_match.group(1)) if wm_match else None
            p_yr=(2025 if p_mo>=8 else 2026) if p_mo else None
            p_T = date(p_yr, p_mo, 1) if p_mo else None
            p_Tend = p_T + timedelta(days=21) if p_T else None
            
            def _m(d, cp=cp, p_T=p_T, p_Tend=p_Tend):
                dl=dt(d.get('d_live'))
                if p_T and not (dl and p_T<=dl<=p_Tend): return False
                if cp!='Other' and canon_product(d['product'])!=cp: return False
                return True
            matched=[d for d in ADV if _m(d)]
            jm_plan.append({'week':pr['week'],'section':pr.get('section',''),'product_type':pr.get('product_type',''),'product':pr['product'],'niche':pr['niche'],
                'planned':pr['planned'],'pic':pr['pic'],'created':pr['created'],'pending':pr.get('pending',''),
                'matched':[{'full':d.get('full') or d['code'],'idea':(d['idea'] or '')[:26],'link':d.get('link') or '','roas':d['roas'],'spend':round(d['spend']),
                    'served':((d.get('days_before_occ') is not None and d['days_before_occ']>=14) if o in OCC_DATE else None)} for d in sorted(matched,key=lambda d:-d['spend'])]})
    acts=act_by_occ.get(o,[]); dated=o in OCC_DATE
    served=sum(1 for d in acts if d.get('days_before_occ') is not None and d['days_before_occ']>=14)
    rows=[{'full':d.get('full') or d['code'],'idea':(d['idea'] or '')[:32],'product':canon_product(d['product']),
           'recipient':(d['recipient'] or '')[:18],'d_ads':d.get('d_ads'),'d_live':d.get('d_live'),
           'days_before':d.get('days_before_occ'),'served':(d.get('days_before_occ') is not None and d['days_before_occ']>=14),'link':d.get('link') or '',
           'spend':round(d['spend']),'profit':round(d['profit']),'roas':d['roas']} for d in sorted(acts,key=lambda d:-d['spend'])]
    quarter = f"Q{(OCC_DATE[o].month-1)//3 + 1} {OCC_DATE[o].year}" if dated else "Ongoing"
    month = OCC_DATE[o].strftime("%B") if dated else "All Year"
    hier.append({'occasion':o,'occ_date':OCC_DATE[o].isoformat() if dated else None,
                 'quarter': quarter, 'month': month,
                 'window':[occ_window(o)[0].isoformat(),occ_window(o)[1].isoformat()] if dated else None,
                 'masterplan':MASTER.get(o,[]), 'jm_timeline':JM_TIMELINE.get(o, []), 'jm_plan':jm_plan,
                 'actuals':{'total':len(acts),'served':served,'served_pct':round(served/len(acts)*100) if (acts and dated) else None,'rows':rows}})
plan['hierarchy']=hier

# ---------- 3. ADS EFFICIENCY ----------
bins=[("ROAS 0–0.5",0,0.5),("0.5–1.0",0.5,1.0),("1.0–1.43 (loss)",1.0,1.43),("1.43–2.0 (win)",1.43,2.0),("2.0–3.0",2.0,3.0),("3.0+",3.0,99)]
roas_bins=[]
for lab,lo,hi in bins:
    grp=[d for d in ADV if lo<=d["roas"]<hi]
    roas_bins.append({"label":lab,"n":len(grp),"spend":round(sum(d["spend"] for d in grp)),"profit":round(sum(d["profit"] for d in grp))})
# spend concentration (Lorenz): cumulative share
sd=sorted(ADV,key=lambda d:-d["spend"]); cum=0; conc=[]
for i,d in enumerate(sd):
    cum+=d["spend"]
    if i in (int(len(sd)*0.01),int(len(sd)*0.05),int(len(sd)*0.1),int(len(sd)*0.25),int(len(sd)*0.5)):
        conc.append({"pct_designs":round((i+1)/len(sd)*100),"pct_spend":round(cum/tot_spend*100)})
untest_spend=sum(d["spend"] for d in ADV if d["bucket"]=="untestable")
byprofit=sorted(ADV,key=lambda d:-d["profit"])
slim=lambda d:{"full":d.get("full") or d["code"],"idea":d["idea"],"link":d.get("link") or "","occasion":d["occasion"],"spend":round(d["spend"]),"profit":round(d["profit"]),"roas":d["roas"],"orders":d["orders"],"bucket":d["bucket"]}
efficiency={"roas_bins":roas_bins,"concentration":conc,"untestable_spend":round(untest_spend),
            "untestable_pct":round(untest_spend/tot_spend*100),"top":[slim(d) for d in byprofit[:15]],"bottom":[slim(d) for d in byprofit[-15:][::-1]]}

# ---------- 4. TIMING (occasion lead-time) ----------
tbins=[("≥4 wks before (in-window)",28,999),("2–4 wks before",14,28),("0–2 wks (late)",0,14),("after occasion (missed)",-999,0)]
timing=[]
occ_adv=[d for d in ADV if d["occasion"] in OCC_DATE and d.get("days_before_occ") is not None]
for lab,lo,hi in tbins:
    grp=[d for d in occ_adv if lo<=d["days_before_occ"]<hi] if lo!=-999 else [d for d in occ_adv if d["days_before_occ"]<0]
    timing.append({"label":lab,"n":len(grp),"spend":round(sum(d["spend"] for d in grp)),"profit":round(sum(d["profit"] for d in grp))})
timing_meta={"occasion_designs":len(occ_adv)}

DASH={"overview":overview,"plan":plan,"efficiency":efficiency,"timing":timing,"timing_meta":timing_meta,
      "scope":"Dec 2025 → now (products live from 2025-12-01)","note":J["note"]}
with open(os.path.join(ROOT,"dash.js"),"w") as fp:
    fp.write("window.DASH="); json.dump(DASH,fp,ensure_ascii=False); fp.write(";")
json.dump(DASH, open(os.path.join(ROOT,"dash.json"),"w"), ensure_ascii=False)
print(f"overview occasions={len(occasions)} | plan rows={len(plan_rows)} weeks={len(weeks)} | served rate={plan['served']['rate']}% (adv {plan['served']['rate_adv']}%)")
print(f"roas bins={[b['n'] for b in roas_bins]} | occasion-timed designs={len(occ_adv)}")
print(f"wrote dash.js ({os.path.getsize(os.path.join(ROOT,'dash.js'))//1024} KB)")