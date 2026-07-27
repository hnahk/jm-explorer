import sqlite3
import json
import os
import re
from datetime import date, timedelta
from openpyxl import load_workbook
import calendar

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
DB_PATH = os.path.join(ROOT, "jm_explorer.db")

def init_db():
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Table: occasions
    c.execute('''
        CREATE TABLE occasions (
            id INTEGER PRIMARY KEY,
            name TEXT UNIQUE,
            occ_date TEXT,
            evergreen BOOLEAN
        )
    ''')

    # Table: designs
    c.execute('''
        CREATE TABLE designs (
            code TEXT PRIMARY KEY,
            full_code TEXT,
            product TEXT,
            niche TEXT,
            recipient TEXT,
            occasion TEXT,
            creator TEXT,
            persona TEXT,
            link TEXT,
            d_live TEXT,
            d_idea TEXT,
            d_media TEXT,
            d_ads TEXT,
            bucket TEXT,
            spend REAL,
            rev REAL,
            profit REAL,
            roas REAL,
            orders INTEGER,
            days_before_occ INTEGER
        )
    ''')

    # Table: campaigns
    c.execute('''
        CREATE TABLE campaigns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            design_code TEXT,
            c_date TEXT,
            spend REAL,
            rev REAL,
            orders INTEGER,
            clicks INTEGER,
            impr INTEGER,
            media_id TEXT,
            FOREIGN KEY(design_code) REFERENCES designs(code)
        )
    ''')

    # Table: plan_rows (from Excel)
    c.execute('''
        CREATE TABLE plan_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            week TEXT,
            section TEXT,
            product_type TEXT,
            product TEXT,
            niche TEXT,
            pic TEXT,
            planned INTEGER,
            created INTEGER,
            pending TEXT
        )
    ''')

    # Table: week_meta (themes etc)
    c.execute('''
        CREATE TABLE week_meta (
            week TEXT PRIMARY KEY,
            theme TEXT,
            start_date TEXT,
            end_date TEXT
        )
    ''')

    # Table: masterplan_idea
    c.execute('''
        CREATE TABLE masterplan_idea (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            occasion TEXT,
            dir TEXT,
            pct REAL,
            searchvol INTEGER,
            rank TEXT
        )
    ''')

    # Table: jm_timeline
    c.execute('''
        CREATE TABLE jm_timeline (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            occasion TEXT,
            week TEXT,
            text TEXT
        )
    ''')

    conn.commit()
    return conn

def mkrange(sn):
    m=re.match(r'(\d+)W(\d+)(?:.*?W?(\d+))?', sn)
    if not m: return None
    mo=int(m.group(1)); w1=int(m.group(2)); w2=int(m.group(3)) if m.group(3) else w1
    yr=2025 if mo>=8 else 2026
    try:
        ld=calendar.monthrange(yr,mo)[1]; return (date(yr,mo,min((w1-1)*7+1,ld)), date(yr,mo,min(w2*7,ld)))
    except: return None

def num(v):
    try: return int(float(v))
    except: return None

def load_data(conn):
    print("Loading data.json...")
    with open(os.path.join(ROOT, "data.json")) as f:
        J = json.load(f)
    
    c = conn.cursor()
    
    # 1. Load Designs
    for d in J["designs"]:
        c.execute('''
            INSERT INTO designs (code, full_code, product, niche, recipient, occasion, creator, persona, link, d_live, d_idea, d_media, d_ads, bucket, spend, rev, profit, roas, orders, days_before_occ)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            d["code"], d.get("full"), d.get("product"), d.get("niche"), d.get("recipient"),
            d.get("occasion"), d.get("creator"), d.get("persona"), d.get("link"),
            d.get("d_live"), d.get("d_idea"), d.get("d_media"), d.get("d_ads"),
            d.get("bucket"), d.get("spend", 0), d.get("rev", 0), d.get("profit", 0),
            d.get("roas", 0), d.get("orders", 0), d.get("days_before_occ")
        ))
        
    # 2. Load Campaigns
    for code, camps in J["campaigns"].items():
        for camp in camps:
            c.execute('''
                INSERT INTO campaigns (design_code, c_date, spend, rev, orders, clicks, impr, media_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                code, camp.get("date"), camp.get("spend", 0), camp.get("rev", 0),
                camp.get("orders", 0), camp.get("clk", 0), camp.get("impr", 0), camp.get("media_id")
            ))

    # 3. Load Occasions (Extract unique occasions from OCC_DATE config in build_dashboard)
    # We will hardcode for now as defined in build_dashboard.py
    OCC_DATE = {'Valentine':'2026-02-14','Easter':'2026-04-05',"Mother's Day":'2026-05-10',
                "Father's Day":'2026-06-21','Graduation':'2026-05-20','4th/Patriotic':'2026-07-04',
                'World Cup':'2026-06-11','Christmas':'2026-12-25','Halloween':'2026-10-31','Back to School':'2026-08-15'}
    for occ, date_str in OCC_DATE.items():
        c.execute('INSERT INTO occasions (name, occ_date, evergreen) VALUES (?, ?, ?)', (occ, date_str, False))
    c.execute('INSERT INTO occasions (name, occ_date, evergreen) VALUES (?, ?, ?)', ('Evergreen', None, True))

    print("Loading Excel Plan (JM Idea's Plan)...")
    try:
        PLAN_XLSX = os.path.join(ROOT, "backend", "sheets", "JM Idea's Plan .xlsx")
        wb=load_workbook(PLAN_XLSX, data_only=True)
        for sn in wb.sheetnames:
            ws=wb[sn]; hr=qi=ci=ni=pi=pti=pici=pendi=None
            for r in range(1,4):
                row=[str(ws.cell(r,col).value or '').lower() for col in range(1,16)]
                if any('idea quantity' in v for v in row):
                    hr=r; qi=next(i+1 for i,v in enumerate(row) if 'idea quantity' in v)
                    ci=next((i+1 for i,v in enumerate(row) if 'created' in v),None)
                    ni=next((i+1 for i,v in enumerate(row) if 'niche' in v),None)
                    pti=next((i+1 for i,v in enumerate(row) if 'product type' in v),None)
                    pici=next((i+1 for i,v in enumerate(row) if v.strip()=='pic'),None)
                    pendi=next((i+1 for i,v in enumerate(row) if 'pending' in v),None)
                    pi=next((i+1 for i,v in enumerate(row) if v.strip()=='product'),None) or next((i+1 for i,v in enumerate(row) if 'product' in v and 'type' not in v),None)
                    break
            if not qi: continue
            theme=''
            for col in range(1,16):
                v=str(ws.cell(hr,col).value or '')
                if any(w in v.lower() for w in ('evergreen','valentine','easter','mother','father','christmas','hobbies','skip','graduation','patriot','independence','summer','world cup')):
                    theme=v.replace('\n',' ').strip()[:50]; break
            wp=wm=0; last_prod=''; last_pt=''; last_sec=''; seci=1 if pti else None
            for r in range(hr+1, ws.max_row+1):
                c1=str(ws.cell(r,1).value or '').strip()
                if c1.upper().startswith('TOTAL'): continue
                q=num(ws.cell(r,qi).value); cr=num(ws.cell(r,ci).value) if ci else None
                if q is None and cr is None: continue
                if seci and c1: last_sec=c1.replace(chr(10),' ').strip()[:44]
                if pti:
                    pt=ws.cell(r,pti).value
                    if pt and str(pt).strip(): last_pt=str(pt).strip().replace(chr(10),' ')[:26]
                prod=ws.cell(r,pi).value if pi else None
                prod=str(prod).strip().replace(chr(10),' ')[:36] if (prod and str(prod).strip()) else ''
                if prod: last_prod=prod
                niche=(str(ws.cell(r,ni).value).strip().replace(chr(10),' ')[:44] if ni and ws.cell(r,ni).value else '')
                pic=(str(ws.cell(r,pici).value).strip()[:12] if pici and ws.cell(r,pici).value else '')
                pend=(str(ws.cell(r,pendi).value).strip().replace(chr(10),' ')[:40] if pendi and ws.cell(r,pendi).value else '')
                
                c.execute('''INSERT INTO plan_rows (week, section, product_type, product, niche, pic, planned, created, pending)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                             (sn, last_sec, last_pt, prod or last_prod, niche, pic, q or 0, cr or 0, pend))
            rg=mkrange(sn)
            c.execute('INSERT INTO week_meta (week, theme, start_date, end_date) VALUES (?, ?, ?, ?)',
                      (sn, theme, rg[0].isoformat() if rg else None, rg[1].isoformat() if rg else None))
        wb.close()
    except Exception as e:
        print(f"Warning: could not load excel plan: {e}")

    print("Loading H1.2026 MASTERPLAN IDEA...")
    try:
        M_XLSX = os.path.join(ROOT, "backend", "sheets", "H1.2026 MASTERPLAN  IDEA.xlsx")
        mwb = load_workbook(M_XLSX, data_only=True)
        for sn,occ_name,prow in [('MD',"Mother's Day",8),('FD',"Father's Day",3)]:
            if sn in mwb.sheetnames:
                ws=mwb[sn]
                for col in range(2,20):
                    dname=str(ws.cell(2,col).value or '').strip()
                    if not dname: continue
                    pv=ws.cell(prow,col).value
                    try: pv=float(pv); pv=round(pv*100) if pv<=1 else round(pv)
                    except: pv=None
                    try: sv=int(float(ws.cell(4,col).value))
                    except: sv=None
                    rank=str(ws.cell(5,col).value or '')[:14]
                    c.execute('INSERT INTO masterplan_idea (occasion, dir, pct, searchvol, rank) VALUES (?, ?, ?, ?, ?)',
                              (occ_name, dname[:40], pv, sv, rank))

        def occ_window(o):
            d=OCC_DATE.get(o); return (date.fromisoformat(d)-timedelta(days=56), date.fromisoformat(d)) if d else None

        if 'TIMELINE' in mwb.sheetnames:
            ws_t = mwb['TIMELINE']; m_str = None
            for col in range(32, 60):
                v5 = str(ws_t.cell(5, col).value or '').strip()
                if v5: m_str = v5
                v6 = str(ws_t.cell(6, col).value or '').strip()
                jm = str(ws_t.cell(15, col).value or '').strip()
                if jm and m_str and v6:
                    m_match = re.search(r'(\d+)', m_str); w_match = re.search(r'(\d+)', v6)
                    if m_match and w_match:
                        mo = int(m_match.group(1)); yr = 2025 if mo >= 8 else 2026
                        w1 = int(w_match.group(1)); ld = calendar.monthrange(yr,mo)[1]
                        ws_date = date(yr,mo,min((w1-1)*7+1,ld))
                        for _o in OCC_DATE:
                            w = occ_window(_o)
                            if w and w[0] <= ws_date <= w[1]:
                                c.execute('INSERT INTO jm_timeline (occasion, week, text) VALUES (?, ?, ?)',
                                          (_o, f'{m_str} {v6}', jm))
        mwb.close()
    except Exception as e:
        print(f"Warning: could not load masterplan: {e}")

    conn.commit()
    print("Done loading JSON and Excel data.")

if __name__ == '__main__':
    conn = init_db()
    load_data(conn)
    conn.close()
